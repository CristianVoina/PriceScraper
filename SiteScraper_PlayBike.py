import os.path
import time
import datetime
import calendar
import sys
import re
from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

dir_path = os.getcwd()
full_path = os.path.join(dir_path,"primulxls.xlsx")

adresa_net_Playbike = sys.argv[1]
adresa_net_BD = sys.argv[2]
adresa_net_B24 = sys.argv[3]

titlu_produs = []
pret_nou_produs = []
pret_vechi_produs = []

def getPret_PLayBike(url):
	content = urlopen(url).read()

	# fisier = open(url, "r", encoding = "utf8")
	# content = fisier.read()

	soup = BeautifulSoup(content, "html.parser")
	#soup = BeautifulSoup(url, "html.parser")

	pret_oferta = soup.find('span', {'class' : 'pret oferta'})
	pret_vechi = soup.find('span', {'class' : 'pret_vechi'})
	titlu = soup.find('h1')

	soup.decompose()
	#fisier.close()
	
	#for preturi in spans
	print("Titlu: " + titlu.text)
	print("Pret Oferta: " + pret_oferta.text)
	print("Pret Vechi: " + pret_vechi.text)
	
	pret_ofer_filt = re.match('[0-9]+', pret_oferta.text)
	pret_vech_filt = re.match('[0-9]+', pret_vechi.text)
	
	return titlu.text, str(pret_ofer_filt.group(0)), str(pret_vech_filt.group(0))
	
def getPret_BD(url):
	content = urlopen(url).read()

	# fisier = open(url, "r", encoding = "utf8")
	# content = fisier.read()

	soup = BeautifulSoup(content, "html.parser")
	#soup = BeautifulSoup(url, "html.parser")

	pret_oferta = soup.find('span', {'class' : 'price'})
	pret_vechi = soup.find('span', {'class' : 'retail-value'})
	titlu = soup.find('div', {'class' : 'title'})
	
	soup.decompose()
	#fisier.close()

	#for preturi in spans
	print("Titlu: " + titlu.text)
	print("Pret Oferta: " + pret_oferta.text)
	print("Pret Vechi: " + pret_vechi.text)
	
	pret_ofer_filt = re.match('[0-9]+', pret_oferta.text)
	pret_vech_filt = re.match('[0-9]+', pret_vechi.text)
	
	return titlu.text, str(pret_ofer_filt.group(0)), str(pret_vech_filt.group(0))
	
def getPret_B24(url):
	content = urlopen(url).read()
	
	# fisier = open(url, "r", encoding = "utf8")
	# content = fisier.read()

	soup = BeautifulSoup(content, "html.parser")

	pret_oferta = soup.find('span', {'class' : 'text-value js-price-value'})
	pret_vechi = soup.find('span', {'class' : 'text-rrp js-text-rrp'})
	titlu = soup.find('h1', {'class' : 'col-md-14 col-lg-14'})
	
	soup.decompose()
	#fisier.close()

	#for preturi in spans
	print("Titlu: " + titlu.text)
	print("Pret Oferta: " + pret_oferta.text)
	print("Pret Vechi: " + pret_vechi.text)
	
	pret_ofer_filt = re.match('[0-9]{3}', pret_oferta.text)
	pret_vech_filt = re.match('.+([0-9]{3}).+', pret_vechi.text)
	
	return titlu.text, str(pret_ofer_filt.group(0)), str(pret_vech_filt.group(1))
	
def xls_appendData(tit, pret_v, pret_n, data):
	if os.path.isfile(full_path):
		wb = load_workbook("primulxls.xlsx")
		ws1 = wb["Preturi"]
		
		coloana = ws1.max_column

		ws1[get_column_letter(coloana+1) + str(1)] = data
		for index, valoare_titlu in enumerate(tit):
			if (valoare_titlu != ws1["A" + str(2 + (3 * index))].value):
				ws1[get_column_letter(coloana+1) + str(2 + (3 * index))] = valoare_titlu
			
		for index, valoare_pretV in enumerate(pret_v):
			ws1[get_column_letter(coloana+1) + str(3 + (3 * index))] = valoare_pretV
			
		for index, valoare_pretN in enumerate(pret_n):
			ws1[get_column_letter(coloana+1) + str(4 + (3 * index))] = valoare_pretN
		
	else:
		wb = Workbook()
		ws1 = wb.active
		ws1.title = "Preturi"	

		ws1["A1"] = data
		for index, valoare_titlu in enumerate(tit):
			ws1["A" + str(2 + (3 * index))] = valoare_titlu 
		
		for index, valoare_pretV in enumerate(pret_v):
			ws1["A" + str(3 + (3 * index))] = valoare_pretV
			
		for index, valoare_pretN in enumerate(pret_n):
			ws1["A" + str(4 + (3 * index))] = valoare_pretN
			
	wb.save("primulxls.xlsx")

	#and (3600 < ((calendar.timegm(time.gmtime())) - (os.path.getmtime(full_path))))
if (os.path.isfile(full_path)) and (3600 < ((calendar.timegm(time.gmtime())) - (os.path.getmtime(full_path)))):
	print("Fisierul deja exista trebuie update")
	print(((calendar.timegm(time.gmtime())) - (os.path.getmtime(full_path))))
	dataCurenta = datetime.datetime.now()
	
	titlu_produs_t, pret_nou_produs_t, pret_vechi_produs_t = getPret_PLayBike(adresa_net_Playbike)
	titlu_produs.append(titlu_produs_t)
	pret_nou_produs.append(pret_nou_produs_t)
	pret_vechi_produs.append(pret_vechi_produs_t)

	titlu_produs_t, pret_nou_produs_t, pret_vechi_produs_t = getPret_BD(adresa_net_BD)
	titlu_produs.append(titlu_produs_t)
	pret_nou_produs.append(pret_nou_produs_t)
	pret_vechi_produs.append(pret_vechi_produs_t)
	
	titlu_produs_t, pret_nou_produs_t, pret_vechi_produs_t = getPret_B24(adresa_net_B24)
	titlu_produs.append(titlu_produs_t)
	pret_nou_produs.append(pret_nou_produs_t)
	pret_vechi_produs.append(pret_vechi_produs_t)
	
	xls_appendData(titlu_produs, pret_vechi_produs, pret_nou_produs, dataCurenta)
	
elif (os.path.isfile(full_path)) == False:

	dataCurenta = datetime.datetime.now()

	titlu_produs_t, pret_nou_produs_t, pret_vechi_produs_t = getPret_PLayBike(adresa_net_Playbike)
	titlu_produs.append(titlu_produs_t)
	pret_nou_produs.append(pret_nou_produs_t)
	pret_vechi_produs.append(pret_vechi_produs_t)

	titlu_produs_t, pret_nou_produs_t, pret_vechi_produs_t = getPret_BD(adresa_net_BD)
	titlu_produs.append(titlu_produs_t)
	pret_nou_produs.append(pret_nou_produs_t)
	pret_vechi_produs.append(pret_vechi_produs_t)
	
	titlu_produs_t, pret_nou_produs_t, pret_vechi_produs_t = getPret_B24(adresa_net_B24)
	titlu_produs.append(titlu_produs_t)
	pret_nou_produs.append(pret_nou_produs_t)
	pret_vechi_produs.append(pret_vechi_produs_t)
	
	xls_appendData(titlu_produs, pret_vechi_produs, pret_nou_produs, dataCurenta)

else:
	print("Fisierul exista dar nu a trecut 1h de la ultimul update")
